#!/usr/bin/env python3
"""
sim_result.py

Simulates placement results for each heat based on heat data.
Uses existing heats generated by heats.py and directly updates each heat tab.
"""

import argparse
import pandas as pd
from openpyxl import load_workbook

from race_utils import get_excel_sheet_names, secure_shuffle


def simulate_and_write_results(file_path: str, heat_sheets: list):
    """
    For each heat sheet, randomize car placements per heat and write back into the same sheet.
    Overwrites values in the existing 'Place' column.

    Args:
        file_path (str): Path to the Excel file to be updated.
        heat_sheets (list): List of sheet names representing heats to simulate.
    """
    wb = load_workbook(file_path)

    for sheet_name in heat_sheets:
        if sheet_name in ("Racers", "Runoff") or sheet_name.endswith("_Rankings"):
            continue

        ws = wb[sheet_name]
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        if "Place" not in df.columns:
            raise ValueError(
                f"Missing 'Place' column in sheet '{sheet_name}'. Please ensure it exists."
            )

        # Simulate results for each heat
        for _, heat_group in df.groupby("Heat"):
            rows = heat_group.index.tolist()
            cars_shuffled = secure_shuffle(heat_group[["Lane", "Car"]].values.tolist())

            for place, (lane, car) in enumerate(cars_shuffled, start=1):
                for idx in rows:
                    if df.at[idx, "Lane"] == lane and df.at[idx, "Car"] == car:
                        df.at[idx, "Place"] = place
                        break

        # Map header column names to worksheet indices
        col_lookup = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)}
        if "Place" not in col_lookup:
            raise ValueError(
                f"'Place' column not found in worksheet '{sheet_name}' header row."
            )

        place_col_idx = col_lookup["Place"]

        for row_idx, value in enumerate(df["Place"], start=2):
            ws.cell(row=row_idx, column=place_col_idx, value=value)

    wb.save(file_path)


def main():
    """
    Main entry point for the simulation script.
    Parses command-line arguments and runs result simulation.
    """
    parser = argparse.ArgumentParser(
        description="Simulate Pinewood Derby heat results."
    )
    parser.add_argument("filename", help="Excel file containing heats to update")
    args = parser.parse_args()

    print(f"Detecting heat sheets in {args.filename}...")
    heat_sheets = get_excel_sheet_names(args.filename)

    print(f"Simulating and writing results to sheets: {heat_sheets}...")
    simulate_and_write_results(args.filename, heat_sheets)

    print("Simulation complete.")


if __name__ == "__main__":
    main()
